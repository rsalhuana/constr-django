# -*- coding: utf-8 -*-
from openpyxl import Workbook
from django.http import HttpResponse
from django.shortcuts import render ,HttpResponse , render_to_response , redirect
from mantenimiento.models import Proyectos , Proveedor , General , GeneralDetalle , Materiales , UnidadMedida , Ubigeo ,Colaborador , Clientes , ColaboradoresCargo
from .forms import MaterialesForm
from .forms import GeneralDetalleForm
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime 
import ast
from django.db.models import Q
from django.template.loader import render_to_string
from django.core.urlresolvers import reverse_lazy
#from django.urls import reverse_lazy
from django.views import generic 
from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.contrib import messages

modal  = 0
idModalNotificar = 0
# Funcion que carga todos los proyectos de ingresos en estado activo 
# @author cvargas
# @date 01/04/2018
def proyectos(request):
	proyectos = Proyectos.objects.filter(b_flagInactivo =0).select_related('id_ubigeo' , 'id_colaborador' , 'vc_codEstado')
	return render(request, 'proyectos/proyectos.html', {'proyectos' : proyectos})

# Funcion que carga todos los clientes
# @author renato ;)
# @date 17/05/2018
def clientes(request):
	clientes = Clientes.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codEstado')
	return render(request, 'clientes/clientes.html', {'clientes' : clientes})

# Funcion que carga los  combos y vista para agregar un proyecto
# @author cvargas
# @date 01/04/2018
def post_nuevoMantenimiento(request):
	documentos = GeneralDetalle.objects.filter(id_general= 8 )
	ubigeo = Ubigeo.objects.filter(vc_codigoubigeo__icontains='1401').order_by('vc_descripcion').distinct()
	colaborador = ColaboradoresCargo.objects.filter(vc_codCargo= 24)
	return render(request, 'proyectos/nuevo.html', {'documentos' : documentos , 'ubigeo' : ubigeo,  'colaborador':colaborador})

# Funcion que carga los  combos y vista para agregar un cliente
# @author renato ;)
# @date 01/04/2018
def nuevoCliente(request):
	documentos = GeneralDetalle.objects.filter(id_general= 8 )
	ubigeo = Ubigeo.objects.filter(vc_codigoubigeo__icontains='1401').order_by('vc_descripcion').distinct()
	return render(request, 'clientes/nuevo.html', {'documentos' : documentos , 'ubigeo' : ubigeo})

# Funcion que permite buscar a un determinado cliente de acuerdo a su numero de documento
# @param  numero de documento del cliente
# @author cvargas
# @date 01/04/2018
@csrf_exempt
def buscarNumDocumento(request):
	documento = request.POST['documento']
	cliente = Clientes.objects.filter(vc_numDocumento= documento ).values('id_cliente', 'vc_razonSocial','vc_codTipoDocumento')
	return HttpResponse( json.dumps(list(cliente)), content_type="application/json")

# Funcion que permite guardar un nuevo cliente
# @param  parametros de cliente
# @author renato ;)
# @date 27/05/2018
def guardarCliente(request):
	ubigeo = None
	estado_cliente = None
	tipo_documento = None
	retenedor = 0

	if request.POST['id_ubigeo']:
		ubigeo = Ubigeo.objects.get(id_ubigeo= request.POST['id_ubigeo'])
	if request.POST['id_estado']:
		estado_cliente = GeneralDetalle.objects.get(id_generalDetalle = request.POST['id_estado'])
	if request.POST['tipo_documento']:
		tipo_documento = GeneralDetalle.objects.get(id_generalDetalle = request.POST['tipo_documento'])
	if 'retenedor' in request.POST:
		retenedor = request.POST['retenedor']
	
	cliente = Clientes(vc_codEstado = estado_cliente,
						vc_razonSocial = request.POST['razon_social'],
						b_retenedor = retenedor,
		                id_ubigeo = ubigeo,
						vc_correo = request.POST['email_cliente'],
						vc_direccion =request.POST['direccion'],
						vc_codTipoDocumento = tipo_documento,
						vc_numDocumento = request.POST['numero_documento'], 
		                vc_contacto=request.POST['contacto'],
						vc_telefono=request.POST['tlf'], 
						b_flaginactivo =0 , 
						dt_crea =datetime.now()	)
	cliente.save()
	global modal
	modal = 1
	return redirect(reverse_lazy('editar_cliente', kwargs={'idCliente': cliente.id_cliente , 'idVista' : 1}))

# Funcion que permite guardar un nuevo proyecto
# @param  parametros de proyecto
# @author cvargas
# @date 01/04/2018
def guardarProyecto(request):
	ubigeo = None
	fechaAprobacion = None
	fechaInicio = None
	fechaCierre = None
	colaborador = None
	if request.POST['fecha_aprobacion']:
		fechaAprobacion =  datetime.strptime(request.POST['fecha_aprobacion'],'%d/%m/%Y').strftime('%Y-%m-%d') 
	if request.POST['fecha_inicio']:
		fechaInicio =  datetime.strptime(request.POST['fecha_inicio'],'%d/%m/%Y').strftime('%Y-%m-%d') 
	if request.POST['fecha_cierre']:
		fechaCierre =  datetime.strptime(request.POST['fecha_cierre'],'%d/%m/%Y').strftime('%Y-%m-%d') 
	estado = GeneralDetalle.objects.get(id_generalDetalle= request.POST['id_estado'])
	if request.POST['id_ubigeo']:
		ubigeo = Ubigeo.objects.get(id_ubigeo= request.POST['id_ubigeo'])
	cliente = Clientes.objects.get(id_cliente= request.POST['id_cliente'])
	if request.POST['id_colaborador']:
		colaborador = Colaborador.objects.get(id_colaborador= request.POST['id_colaborador'])
	proyecto = Proyectos(id_cliente=cliente,vc_codEstado =estado ,
		                 vc_nombreProyecto =request.POST['nombre_proyecto'] , id_ubigeo = ubigeo ,vc_direccion =request.POST['direccion'],
		                 dt_fecAprobacion = fechaAprobacion,dt_fecInicio =fechaInicio,dt_fecCierre =fechaCierre, id_colaborador = colaborador,
		                 vc_nomContacto=request.POST['contacto'],vc_telfContacto=request.POST['tlf'] , b_flagInactivo =0 , dt_crea =datetime.now() )
	proyecto.save()
	global modal
	modal = 1
	return redirect(reverse_lazy('editar_proyecto', kwargs={'idProyecto': proyecto.id_proyecto , 'idVista' : 1}))

# Funcion que consulta los detalles del proyecto seleccionado 
# @param  id del proyecto
# @param  id de la vista de la opcion seleccionada, para conocer si se selecciono la opcion de consulta o de edicion 
# @author cvargas
# @date 02/04/2018
def editarProyecto(request , idProyecto , idVista):
	idModal = 0
	global modal 
	if modal == 1:
		idModal = modal
		modal = 0
	if modal ==2:
		idModal = modal
		modal = 0
	documentos = GeneralDetalle.objects.filter(id_general= 8 )
	ubigeo = Ubigeo.objects.filter(vc_codigoubigeo__icontains='1401').order_by('vc_descripcion').distinct()
	estado =GeneralDetalle.objects.filter(id_general= 7 )
	colaborador = Colaborador.objects.all()
	proyecto = Proyectos.objects.get(id_proyecto = idProyecto)
	return render(request, 'proyectos/editar.html', {'idVista':idVista,'estado':estado,'proyecto':proyecto, 'documentos' : documentos , 'ubigeo' : ubigeo,  'colaborador':colaborador , 'idModal':idModal})

# Funcion que consulta los detalles del cliente seleccionado 
# @param  id del cliente
# @param  id de la vista de la opcion seleccionada, para conocer si se selecciono la opcion de consulta o de edicion 
# @author renato ;)
# @date 25/05/2018
def editarCliente(request , idCliente , idVista):
	idModal = 0
	global modal 
	if modal == 1:
		idModal = modal
		modal = 0
	if modal ==2:
		idModal = modal
		modal = 0
	documentos = GeneralDetalle.objects.filter(id_general= 8 )
	ubigeo = Ubigeo.objects.filter(vc_codigoubigeo__icontains='1401').order_by('vc_descripcion').distinct()
	estado =GeneralDetalle.objects.filter(id_general= 1 )
	
	cliente = Clientes.objects.get(id_cliente = idCliente)
	if cliente.b_retenedor == True :
		cliente.b_retenedor = 1
	else :
		cliente.b_retenedor = 0
	return render(request, 'clientes/editar.html', {'idVista':idVista,'estado':estado,'cliente':cliente, 'documentos' : documentos , 'ubigeo' : ubigeo, 'idModal':idModal})

# Funcion que carga todos los colaboradores
# @author renato ;)
# @date 01/06/2018
def colaboradores(request):
	colaboradores = Colaborador.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codestado')
	return render(request, 'colaboradores/colaboradores.html', {'colaboradores' : colaboradores})

# Funcion que carga los  combos y vista para agregar un colaborador
# @author renato ;)
# @date 01/06/2018
def nuevoColaborador(request):
	ubigeo = Ubigeo.objects.filter(vc_codigoubigeo__icontains='1401').order_by('vc_descripcion').distinct()
	return render(request, 'colaboradores/nuevo.html', {'ubigeo' : ubigeo})

# Funcion que permite guardar un nuevo colaborador
# @param  parametros de colaborador
# @author renato ;)
# @date 01/06/2018
def guardarColaborador(request):
	ubigeo = None
	estado = None
	
	if request.POST['id_ubigeo']:
		ubigeo = Ubigeo.objects.get(id_ubigeo= request.POST['id_ubigeo'])
	if request.POST['id_estado']:
		estado = GeneralDetalle.objects.get(id_generalDetalle = request.POST['id_estado'])
	
	colaborador = Colaborador(vc_codestado = estado,
						vc_razonsocial = request.POST['razon_social'],
						id_ubigeo = ubigeo,
						vc_direccion =request.POST['direccion'],
						vc_ruccolaborador = request.POST['numero_documento'], 
						vc_contacto=request.POST['contacto'],
						vc_telfcontacto=request.POST['telfcontacto'], 
						vc_telfempresa=request.POST['telfempresa'], 
						b_flaginactivo = 0 ,  
						dt_crea =datetime.now()	)
	colaborador.save()
	global modal
	modal = 1
	return redirect(reverse_lazy('editar_colaborador', kwargs={'idColaborador': colaborador.id_colaborador , 'idVista' : 1}))


# Funcion que consulta los detalles del colaborador seleccionado 
# @param  id del colaborador
# @param  id de la vista de la opcion seleccionada, para conocer si se selecciono la opcion de consulta o de edicion 
# @author renato ;)
# @date 01/06/2018
def editarColaborador(request , idColaborador , idVista):
	idModal = 0
	global modal 
	if modal == 1:
		idModal = modal
		modal = 0
	if modal ==2:
		idModal = modal
		modal = 0
	
	ubigeo = Ubigeo.objects.filter(vc_codigoubigeo__icontains='1401').order_by('vc_descripcion').distinct()
	estado =GeneralDetalle.objects.filter(id_general= 1 )
	colaborador = Colaborador.objects.get(id_colaborador = idColaborador)
	
	return render(request, 'colaboradores/editar.html', {'idVista':idVista,'estado':estado,'colaborador':colaborador, 'ubigeo' : ubigeo, 'idModal':idModal})

# Funcion que actualiza los campos del colaborador seleccionado 
# @param  id del colaborador
# @param  detalles del colaborador
# @author renato ;)
# @date 01/06/2018
def guardarEditarColaborador(request):
	ubigeo = None
	estado = None
	
	if request.POST['id_ubigeo']:
		ubigeo = Ubigeo.objects.get(id_ubigeo= request.POST['id_ubigeo'])
	if request.POST['id_estado']:
		estado = GeneralDetalle.objects.get(id_generalDetalle = request.POST['id_estado'])
	
	colaborador = Colaborador.objects.filter(id_colaborador=request.POST['id_colaborador']).update(
									vc_codestado = estado,
									vc_razonsocial = request.POST['razon_social'],
									id_ubigeo = ubigeo,
									vc_direccion =request.POST['direccion'],
									vc_ruccolaborador = request.POST['numero_documento'], 
									vc_contacto=request.POST['contacto'],
									vc_telfcontacto=request.POST['telfcontacto'], 
									vc_telfempresa=request.POST['telfempresa'], 
									b_flaginactivo = 0 , 
									dt_edita =datetime.now())
	global modal
	modal = 2
	return redirect(reverse_lazy('editar_colaborador', kwargs={'idColaborador': request.POST['id_colaborador'] , 'idVista' : 1}))

# Funcion para exportar en formato excel los colaboradores activos de mantenimiento
# @author renato ;)
# @date 01/06/2018
@csrf_exempt
def exportarColaboradores(request):
	colaboradores = Colaborador.objects.filter(b_flaginactivo =0).select_related('id_ubigeo' , 'vc_codestado')
	template_name = "colaboradores/exportar.html"
	response = render_to_response(template_name, {'colaboradores': colaboradores})
	filename = "colaboradores.xls"
	response['Content-Disposition'] = 'attachment; filename='+filename
	response['Content-Type'] = 'application/vnd.ms-excel; charset=utf-16'
	return response

@csrf_exempt
def activeColaborador(request , idColaborador):
	Colaborador.objects.filter(id_colaborador=idColaborador).update(vc_codestado=1)
	colaboradores = Colaborador.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codestado')
	html = render_to_string('colaboradores/tabla_colaboradores.html', {'colaboradores' : colaboradores})
	return HttpResponse(html)

@csrf_exempt
def desactiveColaborador(request , idColaborador ):
	Colaborador.objects.filter(id_colaborador=idColaborador).update(vc_codestado=2)
	colaboradores = Colaborador.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codestado')
	html = render_to_string('colaboradores/tabla_colaboradores.html', {'colaboradores' : colaboradores})
	return HttpResponse(html)

# Funcion que carga todos los proveedores
# @author renato ;)
# @date 31/05/2018
def proveedores(request):
	proveedores = Proveedor.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codestado')
	return render(request, 'proveedores/proveedores.html', {'proveedores' : proveedores})
	
# Funcion que carga los  combos y vista para agregar un proveedor
# @author renato ;)
# @date 31/05/2018
def nuevoProveedor(request):
	ubigeo = Ubigeo.objects.filter(vc_codigoubigeo__icontains='1401').order_by('vc_descripcion').distinct()
	return render(request, 'proveedores/nuevo.html', {'ubigeo' : ubigeo})

# Funcion que permite guardar un nuevo proveedor
# @param  parametros de proveedor
# @author renato ;)
# @date 31/05/2018
def guardarProveedor(request):
	ubigeo = None
	estado = None
	
	if request.POST['id_ubigeo']:
		ubigeo = Ubigeo.objects.get(id_ubigeo= request.POST['id_ubigeo'])
	if request.POST['id_estado']:
		estado = GeneralDetalle.objects.get(id_generalDetalle = request.POST['id_estado'])
	
	proveedor = Proveedor(vc_codestado = estado,
						vc_razonsocial = request.POST['razon_social'],
						id_ubigeo = ubigeo,
						vc_direccion =request.POST['direccion'],
						vc_rucproveedor = request.POST['numero_documento'], 
						vc_contacto=request.POST['contacto'],
						vc_telfcontacto=request.POST['telfcontacto'], 
						vc_telfempresa=request.POST['telfempresa'], 
						b_flaginactivo = 0 ,  
						dt_crea =datetime.now()	)
	proveedor.save()
	global modal
	modal = 1
	return redirect(reverse_lazy('editar_proveedor', kwargs={'idProveedor': proveedor.id_proveedor , 'idVista' : 1}))


# Funcion que consulta los detalles del proveedor seleccionado 
# @param  id del proveedor
# @param  id de la vista de la opcion seleccionada, para conocer si se selecciono la opcion de consulta o de edicion 
# @author renato ;)
# @date 31/05/2018
def editarProveedor(request , idProveedor , idVista):
	idModal = 0
	global modal 
	if modal == 1:
		idModal = modal
		modal = 0
	if modal ==2:
		idModal = modal
		modal = 0
	
	ubigeo = Ubigeo.objects.filter(vc_codigoubigeo__icontains='1401').order_by('vc_descripcion').distinct()
	estado =GeneralDetalle.objects.filter(id_general= 1 )
	proveedor = Proveedor.objects.get(id_proveedor = idProveedor)
	
	return render(request, 'proveedores/editar.html', {'idVista':idVista,'estado':estado,'proveedor':proveedor, 'ubigeo' : ubigeo, 'idModal':idModal})

# Funcion que actualiza los campos del proveedor seleccionado 
# @param  id del proveedor
# @param  detalles del proveedor
# @author renato ;)
# @date 31/05/2018
def guardarEditarProveedor(request):
	ubigeo = None
	estado = None
	
	if request.POST['id_ubigeo']:
		ubigeo = Ubigeo.objects.get(id_ubigeo= request.POST['id_ubigeo'])
	if request.POST['id_estado']:
		estado = GeneralDetalle.objects.get(id_generalDetalle = request.POST['id_estado'])
	
	proveedor = Proveedor.objects.filter(id_proveedor=request.POST['id_proveedor']).update(
									vc_codestado = estado,
									vc_razonsocial = request.POST['razon_social'],
									id_ubigeo = ubigeo,
									vc_direccion =request.POST['direccion'],
									vc_rucproveedor = request.POST['numero_documento'], 
									vc_contacto=request.POST['contacto'],
									vc_telfcontacto=request.POST['telfcontacto'], 
									vc_telfempresa=request.POST['telfempresa'], 
									b_flaginactivo = 0 , 
									dt_edita =datetime.now())
	global modal
	modal = 2
	return redirect(reverse_lazy('editar_proveedor', kwargs={'idProveedor': request.POST['id_proveedor'] , 'idVista' : 1}))

# Funcion para exportar en formato excel los proveedores activos de mantenimiento
# @author renato ;)
# @date 31/05/2018
@csrf_exempt
def exportarProveedores(request):
	proveedores = Proveedor.objects.filter(b_flaginactivo =0).select_related('id_ubigeo' , 'vc_codestado')
	template_name = "proveedores/exportar.html"
	response = render_to_response(template_name, {'proveedores': proveedores})
	filename = "proveedores.xls"
	response['Content-Disposition'] = 'attachment; filename='+filename
	response['Content-Type'] = 'application/vnd.ms-excel; charset=utf-16'
	return response

@csrf_exempt
def activeProveedor(request , idProveedor):
	Proveedor.objects.filter(id_proveedor=idProveedor).update(vc_codestado=1)
	proveedores = Proveedor.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codestado')
	html = render_to_string('proveedores/tabla_proveedores.html', {'proveedores' : proveedores})
	return HttpResponse(html)

@csrf_exempt
def desactiveProveedor(request , idProveedor ):
	Proveedor.objects.filter(id_proveedor=idProveedor).update(vc_codestado=2)
	proveedores = Proveedor.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codestado')
	html = render_to_string('proveedores/tabla_proveedores.html', {'proveedores' : proveedores})
	return HttpResponse(html)

# Funcion que actualiza los campos del proyecto seleccionado 
# @param  id del proyecto
# @param  detalles del proyecto
# @author cvargas
# @date 02/04/2018
def guardarEditarProyecto(request):
	ubigeo = None
	fechaAprobacion = None
	fechaInicio = None
	fechaCierre = None
	colaborador = None
	if  request.POST['fecha_aprobacion']:
		fechaAprobacion =  datetime.strptime(request.POST['fecha_aprobacion'],'%d/%m/%Y').strftime('%Y-%m-%d') 
	if request.POST['fecha_inicio']:
		fechaInicio =  datetime.strptime(request.POST['fecha_inicio'],'%d/%m/%Y').strftime('%Y-%m-%d') 
	if request.POST['fecha_cierre']:
		fechaCierre =  datetime.strptime(request.POST['fecha_cierre'],'%d/%m/%Y').strftime('%Y-%m-%d') 
	estado = GeneralDetalle.objects.get(id_generalDetalle= request.POST['id_estado'])
	if request.POST['id_ubigeo']:
		ubigeo = Ubigeo.objects.get(id_ubigeo= request.POST['id_ubigeo'])
	if request.POST['id_colaborador']:
		colaborador = Colaborador.objects.get(id_colaborador= request.POST['id_colaborador'])
	cliente = Clientes.objects.get(id_cliente= request.POST['id_cliente'])
	proyecto = Proyectos.objects.filter(id_proyecto=request.POST['id_proyecto']).update(vc_codEstado =estado , id_cliente = cliente,
		                 vc_nombreProyecto =request.POST['nombre_proyecto'] , id_ubigeo = ubigeo ,vc_direccion =request.POST['direccion'],
		                 dt_fecAprobacion = fechaAprobacion,dt_fecInicio =fechaInicio,dt_fecCierre =fechaCierre, id_colaborador = colaborador,
		                 vc_nomContacto=request.POST['contacto'],vc_telfContacto=request.POST['tlf'] , b_flagInactivo =0 , 
		                 dt_edita =datetime.now())
	global modal
	modal = 2
	return redirect(reverse_lazy('editar_proyecto', kwargs={'idProyecto': request.POST['id_proyecto'] , 'idVista' : 1}))

# Funcion que actualiza los campos del cliente seleccionado 
# @param  id del cliente
# @param  detalles del cliente
# @author renato ;)
# @date 27/05/2018
def guardarEditarCliente(request):
	ubigeo = None
	estado_cliente = None
	tipo_documento = None
	retenedor = 0

	if request.POST['id_ubigeo']:
		ubigeo = Ubigeo.objects.get(id_ubigeo= request.POST['id_ubigeo'])
	if request.POST['id_estado']:
		estado_cliente = GeneralDetalle.objects.get(id_generalDetalle = request.POST['id_estado'])
	if request.POST['tipo_documento']:
		tipo_documento = GeneralDetalle.objects.get(id_generalDetalle = request.POST['tipo_documento'])
	if 'retenedor' in request.POST:
		retenedor = request.POST['retenedor']

	cliente = Clientes.objects.filter(id_cliente=request.POST['id_cliente']).update(
									vc_codEstado = estado_cliente,
									vc_razonSocial = request.POST['razon_social'],
									id_ubigeo = ubigeo,
									b_retenedor = retenedor,
									vc_correo = request.POST['email_cliente'],
									vc_direccion =request.POST['direccion'],
									vc_codTipoDocumento = tipo_documento,
									vc_numDocumento = request.POST['numero_documento'], 
									vc_contacto=request.POST['contacto'],
									vc_telefono=request.POST['tlf'], 
									b_flaginactivo = 0 , 
									dt_edita =datetime.now())
	global modal
	modal = 2
	return redirect(reverse_lazy('editar_cliente', kwargs={'idCliente': request.POST['id_cliente'] , 'idVista' : 1}))


# Funcion para actualizar el estado del proyecto seleccionado a Registrado o Anulado
# @param  id del proyecto a actualizar
# @param  id del estado en el que se encuentra el proyecto
# @author cvargas
# @date 03/04/2018
@csrf_exempt
def estatusProyecto(request):
	idEstado =request.POST['idEstado']
	codigo = 20
	if idEstado != '16':
		codigo = 16
	proyecto = Proyectos.objects.filter(id_proyecto=request.POST['id']).update(vc_codEstado=codigo)
	proyectos = Proyectos.objects.filter(b_flagInactivo =0).select_related('id_ubigeo' , 'id_colaborador' , 'vc_codEstado')
	html = render_to_string('proyectos/tabla_proyectos.html', {'proyectos' : proyectos})
	return HttpResponse(html)

# Funcion para exportar en formato excel los proyectos activos de mantenimiento
# @author cvargas
# @date 03/04/2018
@csrf_exempt
def exportarProyectos(request):
	proyectos = Proyectos.objects.filter(b_flagInactivo =0).select_related('id_ubigeo' , 'id_colaborador' , 'vc_codEstado')
	template_name = "proyectos/exportar.html"
	response = render_to_response(template_name, {'proyectos': proyectos})
	filename = "proyectos.xls"
	response['Content-Disposition'] = 'attachment; filename='+filename
	response['Content-Type'] = 'application/vnd.ms-excel; charset=utf-16'
	return response

# Funcion para exportar en formato excel los proyectos activos de mantenimiento
# @author cvargas
# @date 03/04/2018
@csrf_exempt
def exportarClientes(request):
	clientes = Clientes.objects.filter(b_flaginactivo =0).select_related('id_ubigeo' , 'vc_codEstado')
	template_name = "clientes/exportar.html"
	response = render_to_response(template_name, {'clientes': clientes})
	filename = "clientes.xls"
	response['Content-Disposition'] = 'attachment; filename='+filename
	response['Content-Type'] = 'application/vnd.ms-excel; charset=utf-16'
	return response

def materiales(request):
	idModal = 0
	global modal 
	if modal == 1:
		idModal = modal
		modal = 0
	materiales = Materiales.objects.filter(b_flagInactivo =0).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	return render(request, 'materiales/materiales.html', {'materiales' : materiales , 'idModal' : modal})

class MaterialesCreateView(generic.edit.CreateView):
	model = Materiales
	form_class = MaterialesForm
	template_name = 'materiales/nuevo.html' 
	success_url = reverse_lazy('materiales:index')

	def get(self, request, * args, ** kwargs):
		form = self.form_class()
		return render(request, self.template_name, {'form': form})

	def post(self, request, * args, ** kwargs):
		form = self.form_class(request.POST)
		if form.is_valid():
			idColor = request.POST.get('id_id_vcodColor')
			if request.POST.get('vc_valor1'):
				general = General.objects.get(id_general = 13)
				generalDetalle = GeneralDetalle(id_general = general , vc_codigo ='001' , vc_valor1 = request.POST.get('vc_valor1'), dt_crea = datetime.now(),
            			vc_usuarioCrea = request.user.username , b_flagInactivo = 0 , vc_ipCrea = request.META.get('REMOTE_ADDR'))
				generalDetalle.save()
				idColor =  generalDetalle.id_generalDetalle
			materiales = form.save(commit=False)
			materiales.dt_crea = datetime.now()
			materiales.vc_usuarioCrea = request.user.username
			materiales.b_flagInactivo = 0
			materiales.nu_stockActual = 0.00
			if idColor != None:
				materiales.id_codColor = GeneralDetalle.objects.get(id_generalDetalle = idColor)
			materiales.id_codEstado = GeneralDetalle.objects.get(id_generalDetalle = 42)
			materiales.id_codSituacion = GeneralDetalle.objects.get( id_generalDetalle = 25)
			materiales.vc_ipCrea = request.META.get('REMOTE_ADDR')
			materiales.save()
			lastId = 	Materiales.objects.latest('id_materiales')
			messages.success(request, 'La información se grabó Satisfactoriamente')
			return redirect(reverse_lazy('editar_material' , kwargs={ 'pk' : lastId.id_materiales , 'tipo' : 1 }))
		return render(request, self.template_name, {'form': form})

class MaterialesUpdateView(generic.edit.UpdateView):
	model = Materiales
	form_class = MaterialesForm
	template_name = 'materiales/nuevo.html' 
	success_url = reverse_lazy('materiales:index')  
	def post(self, request, * args, ** kwargs):
		self.object = self.get_object()
		form = self.form_class(request.POST, instance=self.object)
		form.vc_usuarioedita = request.user
		form.dt_edita = datetime.now()
		if form.is_valid(): 
			materiales = form.save(commit=False)
			materiales.b_flagInactivo = 0
			materiales.vc_ipEdita = request.META.get('REMOTE_ADDR')
			materiales.dt_edita = datetime.now()
			materiales.vc_usuarioEdita = request.user.username
			materiales.save()
			messages.success(request, 'Registro actualizado satisfactoriamente.')
			return render(request, self.template_name, {'form': form})
		return render(request, self.template_name, {'form': form })

@csrf_exempt
def buscarMateriales(request):
	edit = 0
	ver = 0
	delete = 0
	if request.POST.get('id_edit'):
		edit = 1
	if request.POST.get('id_delete'):
		delete = 1
	if request.POST.get('id_ver'):
		ver = 1
	search = request.POST.get('busqueda')
	materiales = Materiales.objects.filter((Q(vc_codigo__icontains=search) | Q(vc_descripcion__icontains=search)
									  | Q(nu_valor__icontains=search)
                                      | Q(id_unidadMedida__vc_descripcionCorta__icontains=search)
                                      | Q(nu_stockActual__icontains=search)
                                      | Q(id_codSituacion__vc_valor1__icontains=search)
                                      | Q(id_codEstado__vc_valor1__icontains=search)
                                      )& Q(b_flagInactivo = 0))
	
	html = render_to_string('materiales/tabla_materiales.html', {'materiales' : materiales , 'edit' : edit , 'ver' : ver , 'delete' : delete})
	return HttpResponse(html)

@csrf_exempt
def exportarMateriales(request , search):
	if search == 0:
		search = None
	materiales = Materiales.objects.filter((Q(vc_codigo__icontains=search) | Q(vc_descripcion__icontains=search)
									  | Q(nu_valor__icontains=search)
                                      | Q(id_unidadMedida__vc_descripcionCorta__icontains=search)
                                      | Q(nu_stockActual__icontains=search)
                                      | Q(id_codSituacion__vc_valor1__icontains=search)
                                      | Q(id_codEstado__vc_valor1__icontains=search)
                                      )& Q(b_flagInactivo = 0))
	wb = Workbook()
	ws = wb.active
	ws['A1'] = 'Código'
	ws['B1'] = 'Descripción'
	ws['C1'] = 'Valor'
	ws['D1'] = 'Unid. Med'
	ws['E1'] = 'Stock Actual'
	ws['F1'] = 'Situación'
	ws['G1'] = 'Estado'
	cont=2
	for material in materiales:
		ws.cell(row=cont,column=1).value = material.vc_codigo
		ws.cell(row=cont,column=2).value = material.vc_descripcion
		ws.cell(row=cont,column=3).value = material.nu_valor
		ws.cell(row=cont,column=4).value = material.id_unidadMedida.vc_descripcionCorta
		ws.cell(row=cont,column=5).value = material.nu_stockActual
		ws.cell(row=cont,column=6).value = material.id_codSituacion.vc_valor1
		ws.cell(row=cont,column=7).value = material.id_codEstado.vc_valor1
		cont = cont + 1
	nombre_archivo ="materiales.xlsx"
	response = HttpResponse(content_type="application/ms-excel")
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response

@csrf_exempt
def activeMaterial(request , idMaterial):
	Materiales.objects.filter(id_materiales=idMaterial).update(id_codEstado=42)
	materiales = Materiales.objects.filter(b_flagInactivo =0).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	html = render_to_string('materiales/tabla_materiales.html', {'materiales' : materiales})
	return HttpResponse(html)

@csrf_exempt
def desactiveMaterial(request , idMaterial ):
	Materiales.objects.filter(id_materiales=idMaterial).update(id_codEstado=43)
	materiales = Materiales.objects.filter(b_flagInactivo =0).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	html = render_to_string('materiales/tabla_materiales.html', {'materiales' : materiales})
	return HttpResponse(html)

@csrf_exempt
def activeCliente(request , idCliente):
	Clientes.objects.filter(id_cliente=idCliente).update(vc_codEstado=1)
	#clientes = Clientes.objects.filter(b_flagInactivo =0).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	clientes = Clientes.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codEstado')
	html = render_to_string('clientes/tabla_clientes.html', {'clientes' : clientes})
	return HttpResponse(html)

@csrf_exempt
def desactiveCliente(request , idCliente ):
	Clientes.objects.filter(id_cliente=idCliente).update(vc_codEstado=2)
	#clientes = Clientees.objects.filter(b_flagInactivo =0).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	clientes = Clientes.objects.filter(b_flaginactivo =0).select_related('id_ubigeo', 'vc_codEstado')
	html = render_to_string('clientes/tabla_clientes.html', {'clientes' : clientes})
	return HttpResponse(html)

def viewNotificar(request):
	idModal = 0
	global idModalNotificar 
	if idModalNotificar == 1:
		idModal = idModalNotificar
		idModalNotificar = 0
	else:
		idModal = 1
	materiales = Materiales.objects.filter(b_flagInactivo =0 , id_codSituacion = 25 , id_codEstado=42).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	toEmail = GeneralDetalle.objects.get(id_generalDetalle = 44 )
	return render(request, 'materiales/notificar.html', {'materiales' : materiales , 'toEmail':toEmail  , 'idNotificar' : idModal})


def notificarMaterial(request):
	idNotificar = request.POST['idNotificar']
	global idModalNotificar
	idModalNotificar = 1
	if idNotificar == '1':
		materiales = Materiales.objects.filter(b_flagInactivo =0 , id_codSituacion = 25 , id_codEstado=42).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
		msg_html = render_to_string('materiales/tabla_notificar.html', {'materiales': materiales})
		msg  = EmailMultiAlternatives(subject=request.POST['id_subject'],
	            body=msg_html,
	            to=[request.POST['id_from']],
	            cc=[request.POST['id_cc']])
		msg.attach_alternative(msg_html, "text/html")
		msg.send()
		fecha = datetime.now()
		year = fecha.year
		var = Materiales.objects.all()
		x = 0
		for v in var:
			print(v.vc_numNotificacion)
			if v.vc_numNotificacion != None:
			 	x = 1
		if x == 0:
			vcodigo =  str(year) + '-00001'
		else:
			codigo = Materiales.objects.filter(~Q(id_codSituacion = 25)).order_by('-vc_numNotificacion').exclude( vc_numNotificacion = None).latest('vc_numNotificacion')
			if codigo.vc_numNotificacion:
				inicial = codigo.vc_numNotificacion[0:4]
				lastCodigo =  codigo.vc_numNotificacion[5:10]
				newCodigo = int(lastCodigo.lstrip('+0')) + 1
				lenCodigo = len(str(newCodigo))
				if lenCodigo == 10:
					vcodigo = str(year) + '-' + newCodigo
				else:
					vcodigo =  str(year) + '-' + str(newCodigo).zfill(5)
			else:
				vcodigo =  str(year) + '-00001'
		for idUpdate in request.POST.getlist('id_notificados'):
			Materiales.objects.filter(id_materiales=idUpdate).update(vc_numNotificacion = vcodigo , id_codSituacion = 27)
		messages.success(request, 'Mensaje enviado con exito')
		return redirect(reverse_lazy('view_notificar'))
	else:
		materiales = Materiales.objects.filter(b_flagInactivo =0 , id_codSituacion = 26 , id_codEstado=42).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
		msg_html = render_to_string('materiales/tabla_notificar.html', {'materiales': materiales})
		msg  = EmailMultiAlternatives(subject=request.POST['id_subject'],
	            body=msg_html,
	            to=[request.POST['id_from']],
	            cc=[request.POST['id_cc']])
		msg.attach_alternative(msg_html, "text/html")
		msg.send()
		for idUpdate in request.POST.getlist('id_notificados'):
			Materiales.objects.filter(id_materiales=idUpdate).update(id_statusAprobado = 1)
		messages.success(request, 'Mensaje enviado con exito')
		return redirect(reverse_lazy('view_notificar_aprobados'))

def viewNotificarAprobados(request):
	idModal = 0
	global idModalNotificar 
	if idModalNotificar == 1:
		idModal = idModalNotificar
		idModalNotificar = 0
	else:
		idModal = 2
	materiales = Materiales.objects.filter(b_flagInactivo =0 , id_codSituacion = 26 , id_codEstado=42 , id_statusAprobado = 0).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	toEmail = GeneralDetalle.objects.get(id_generalDetalle = 45 )
	return render(request, 'materiales/notificar.html', {'materiales' : materiales , 'toEmail':toEmail , 'idNotificar' : idModal})


def viewAprobar(request):
	materiales = Materiales.objects.filter(b_flagInactivo =0 , id_codSituacion = 27 , id_codEstado=42).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	return render(request, 'materiales/aprobar.html', {'materiales' : materiales })

def verificar(request , idMaterial):
	data = Materiales.objects.get(id_materiales = idMaterial)
	materiales = Materiales.objects.filter(b_flagInactivo =0 , id_codSituacion = 26 , id_codEstado=42).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	return render(request, 'materiales/verificar.html', {'data':data ,  'materiales' : materiales})

@csrf_exempt
def buscarAprobados(request):
	search = request.POST.get('busqueda')
	materiales = Materiales.objects.filter((Q(vc_codigo__icontains=search) 
									  | Q(id_codFamilia__vc_valor1__icontains=search)
									  | Q(id_codClasificacion__vc_valor1__icontains=search)
									  | Q(vc_descripcion__icontains=search)
									  | Q(id_codColor__vc_valor1__icontains=search)
									  | Q(nu_valor__icontains=search)
                                      | Q(id_unidadMedida__vc_descripcionCorta__icontains=search)
                                      | Q(id_codSituacion__vc_valor1__icontains=search)
                                      | Q(id_codEstado__vc_valor1__icontains=search)
                                      )& Q(b_flagInactivo = 0) & Q(id_codSituacion = 26) & Q(id_codEstado=42))
	
	html = render_to_string('materiales/tabla_verificar.html', {'materiales' : materiales})
	return HttpResponse(html)

@csrf_exempt
def desaprobarMaterial(request , idMaterial):
	observacion = request.POST.get('observacion')
	Materiales.objects.filter(id_materiales=idMaterial).update(vc_observacionDesaprobar=observacion , id_codSituacion = 28)
	materiales = Materiales.objects.filter(b_flagInactivo =0 , id_codSituacion = 25 , id_codEstado=42).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	return render(request, 'materiales/tabla_aprobar.html', {'materiales' : materiales})


@csrf_exempt
def generarCodigo(request):
	idMaterial = request.POST.get('id')
	material = Materiales.objects.get(id_materiales = idMaterial)
	tipoGasto = material.id_codTipoGasto.id_generalDetalle
	situacion = material.id_codSituacion.id_generalDetalle
	materiales = Materiales.objects.filter(id_codTipoGasto=tipoGasto).order_by('-vc_codigo').exclude( vc_codigo = None).latest('vc_codigo')
	if materiales.vc_codigo:
		inicial = materiales.vc_codigo[0:2]
		lastCodigo =  materiales.vc_codigo[2:7]
		newCodigo = int(lastCodigo.lstrip('+0')) + 1
		lenCodigo = len(str(newCodigo))
		if lenCodigo == 5:
			codigo = inicial + newCodigo
		else:
			codigo =  inicial + str(newCodigo).zfill(5)
	else:
		detalle = GeneralDetalle.objects.get(id_generalDetalle = tipoGasto)
		codigo = detalle.vc_valor2 + '00001'
	return HttpResponse( json.dumps(codigo), content_type="application/json")

@csrf_exempt
def aprobarMaterial(request , idMaterial):
	codigo = request.POST.get('codigo')
	Materiales.objects.filter(id_materiales=idMaterial).update(vc_codigo=codigo , id_codSituacion = 26 , id_codEstado = 42 , id_statusAprobado = 0)
	materiales = Materiales.objects.filter(b_flagInactivo =0 , id_codSituacion = 27 , id_codEstado=42).select_related('id_codSituacion' , 'id_codEstado' , 'id_unidadMedida')
	return render(request, 'materiales/tabla_aprobar.html', {'materiales' : materiales})
